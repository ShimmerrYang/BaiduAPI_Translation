# -*- coding: utf-8 -*-

import requests
import random
import json
from hashlib import md5
import openpyxl
import time

# name = input('请输入Excel的文件名：')
qianzhui = 'demo'
name = qianzhui+'.xlsx'

# 一次提交可翻译的行数
limit = 4

# User: YzH
appid = '20210325000742262'
appkey = 'nBZkYdALLeoNfFhFxcfT'

# For list of language codes, please refer to `https://api.fanyi.baidu.com/doc/21`
from_lang = 'en'
to_lang = 'zh'

# -------------------------------参数设置完成-------------------------------

endpoint = 'http://api.fanyi.baidu.com'
path = '/api/trans/vip/translate'
url = endpoint + path


# Generate salt and sign
def make_md5(s, encoding='utf-8'):
    return md5(s.encode(encoding)).hexdigest()


wb = openpyxl.load_workbook(name)
ws = wb.active

if ws['C1'].value != '标题翻译':
    print('格式错误，请检查')
    exit()

# 确定提交到循环的次数number
yushu = (ws.max_row-1) % limit
if yushu == 0:
    number = int((ws.max_row-1)/limit)
else:
    number = (ws.max_row-1)//limit

f = 0  # 新的就是0，断点续跑为读取次数减一
# 开始循环提交翻译
for step in range(f, number):

    query = ''
    for i in range(2+step*limit, 2+(step+1)*limit):  # 第一次，2+0*7，第2行开始
        if ws['C'+str(i)].value != None:
            query = query+ws['C'+str(i)].value+'\n'
            query = query+ws['D'+str(i)].value+'\n'
        else:
            query = query + '*\n' + '*\n'

    print('第{}次读取完成，开始提交数据...'.format(step+1))
    print(query)

    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)

    # Build request
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}

    # Send request
    r = requests.post(url, params=payload, headers=headers)
    result = r.json()

    print(json.dumps(result, indent=4, ensure_ascii=False))  # 对result可视化显示

    # Save in sheet
    index = 0
    for i in range(2+step*limit, 2+(step+1)*limit):  # 第一次，2+0*7，第2行开始
        ws['C'+str(i)].value = result['trans_result'][index]['dst']
        ws['D'+str(i)].value = result['trans_result'][index+1]['dst']
        index = index+2

    wb.save(qianzhui+'_翻译版.xlsx')
    time.sleep(2.6)

# 如果有余数再提交最后一遍
if yushu != 0:
    query = ''
    for i in range(2 + number * limit, 2 + number * limit + yushu):
        query = query + ws['C' + str(i)].value + '\n'
        query = query + ws['D' + str(i)].value + '\n'

    print('第{}次读取完成，开始提交数据...'.format(number+1))
    print(query)

    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)

    # Build request
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}

    # Send request
    r = requests.post(url, params=payload, headers=headers)
    result = r.json()

    print(json.dumps(result, indent=4, ensure_ascii=False))  # 对result可视化显示

    # Save in sheet
    index = 0
    for i in range(2 + number * limit, 2 + number * limit + yushu):
        ws['C' + str(i)].value = result['trans_result'][index]['dst']
        ws['D' + str(i)].value = result['trans_result'][index + 1]['dst']
        index = index + 2

    wb.save(qianzhui+'_翻译版.xlsx')

print('翻译完成！！！')
